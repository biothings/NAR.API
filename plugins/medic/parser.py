"""
MeDIC parser — Medicines, Diseases, Indications, and Contraindications

Parses 3 files from GitHub releases (everycure-org/matrix-*):
  - drugList.tsv           — drug metadata, cross-identifiers, ATC codes
  - indicationList.xlsx    — curated government-regulatory drug indications
  - contraindicationList.xlsx — curated government-regulatory contraindications

Primary key: drug CURIE (e.g. CHEBI:8327, UNII:N0A21N6RAU) from drugList.tsv.
One document per drug — indications and contraindications nested as lists.
Target API: MyChem.info (CHEBI IDs map natively; InChIKey in xrefs when available)
"""

import os
import csv
import ast
from collections import defaultdict
from biothings.utils.dataload import dict_sweep, unlist

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clean_val(value):
    """Return None for empty/NaN, else stripped string."""
    if value is None:
        return None
    v = str(value).strip()
    if v in ("", "nan", "None", "NaN"):
        return None
    return v


def _to_bool(value):
    """Parse boolean-like values."""
    if value is None:
        return None
    v = str(value).strip().upper()
    if v in ("TRUE", "1", "YES"):
        return True
    if v in ("FALSE", "0", "NO"):
        return False
    return None


def _parse_alternate_ids(raw):
    """
    Parse the alternate_ids column from drugList.tsv.
    Format: Python list literal string e.g. "['INCHIKEY:X', 'RXCUI:123']"
    Returns dict: {prefix: [value, ...]}
    """
    if not raw or raw.strip() in ("", "nan", "None"):
        return {}
    try:
        items = ast.literal_eval(raw.strip())
        if not isinstance(items, list):
            return {}
    except (ValueError, SyntaxError):
        # Fallback: split on comma-space
        items = [i.strip().strip("'[]") for i in raw.split(",") if i.strip()]

    result = defaultdict(list)
    for item in items:
        item = item.strip()
        if ":" in item:
            prefix, val = item.split(":", 1)
            prefix = prefix.strip().upper()
            val = val.strip()
            if val and prefix:
                result[prefix].append(val)
    return dict(result)


def _extract_inchikey(alt_ids_dict):
    """Extract InChIKey from parsed alternate_ids dict."""
    keys = alt_ids_dict.get("INCHIKEY", [])
    return keys[0] if keys else None


# ---------------------------------------------------------------------------
# Phase 1: Load drug list
# ---------------------------------------------------------------------------

def _load_drug_list(data_folder):
    """
    Load drugList.tsv. Returns dict: curie -> drug metadata.
    """
    drugs = {}
    fpath = os.path.join(data_folder, "drugList.tsv")
    if not os.path.exists(fpath):
        return drugs

    with open(fpath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            curie = _clean_val(row.get("curie"))
            if not curie or curie == "curie":
                continue

            alt_ids = _parse_alternate_ids(row.get("alternate_ids", ""))
            inchikey = _extract_inchikey(alt_ids)

            # Build xrefs dict from parsed alternate_ids
            xrefs = {}
            for prefix, vals in alt_ids.items():
                prefix_lower = prefix.lower()
                if prefix_lower in ("inchikey",):
                    continue  # stored separately
                # Normalize prefix names
                key = {
                    "pubchem.compound": "pubchem_cid",
                    "chembl.compound": "chembl",
                    "drugbank": "drugbank",
                    "rxcui": "rxcui",
                    "mesh": "mesh",
                    "unii": "unii",
                    "umls": "umls",
                    "chebi": "chebi",
                    "cas": "cas",
                    "gtopdb": "gtopdb",
                    "hmdb": "hmdb",
                }.get(prefix_lower, prefix_lower)
                deduped = list(dict.fromkeys(vals))  # deduplicate while preserving order
                xrefs[key] = deduped if len(deduped) > 1 else deduped[0]

            # ATC codes — stored as Python list literal e.g. "['A10BX02', 'A10BX03']"
            atc_codes_raw = _clean_val(row.get("atc_codes"))
            if atc_codes_raw:
                try:
                    atc_parsed = ast.literal_eval(atc_codes_raw)
                    atc_codes = list(dict.fromkeys(atc_parsed)) if isinstance(atc_parsed, list) else [str(atc_parsed)]
                except (ValueError, SyntaxError):
                    atc_codes = [a.strip() for a in atc_codes_raw.strip("[]'\"").split(",") if a.strip()]
            else:
                atc_codes = None

            # Approval status flags
            approved_usa = _to_bool(row.get("approved_usa"))
            approved_eu = _to_bool(row.get("approved_eu"))
            approved_japan = _to_bool(row.get("approved_japan"))

            # Classification flags (boolean columns)
            flags = {}
            for col in [
                "is_radioisotope_or_diagnostic_agent", "is_allergen",
                "is_metallic_salt", "is_steroid", "is_antimicrobial",
                "is_chemotherapy", "is_glucose_regulator",
                "is_vaccine_or_antigen", "is_no_therapeutic_value",
            ]:
                v = _to_bool(row.get(col))
                if v:  # only include True flags
                    flags[col] = v

            doc = {
                "drug_id": curie,
                "name": _clean_val(row.get("drug_name")) or _clean_val(row.get("curie_label")),
                "combination_therapy": _to_bool(row.get("combination_therapy")),
                "atc_codes": atc_codes,
                "atc_main": _clean_val(row.get("atc_main")),
                "approval_status": {
                    "usa": approved_usa,
                    "eu": approved_eu,
                    "japan": approved_japan,
                },
                "xrefs": xrefs if xrefs else None,
                "inchikey": inchikey,
                "smiles": _clean_val(row.get("smiles")),
            }
            if flags:
                doc["classification_flags"] = flags

            drugs[curie] = doc

    return drugs


# ---------------------------------------------------------------------------
# Phase 2: Load indication list
# ---------------------------------------------------------------------------

def _load_indications(data_folder):
    """
    Load indicationList.xlsx. Returns dict: drug_curie -> list of indication records.
    """
    indications = defaultdict(list)
    fpath = os.path.join(data_folder, "indicationList.xlsx")
    if not os.path.exists(fpath):
        return indications

    if HAS_PANDAS:
        df = pd.read_excel(fpath, engine="openpyxl")
        df = df.where(pd.notnull(df), None)
        for _, row in df.iterrows():
            drug_id = _clean_val(row.get("final normalized drug id"))
            dis_id = _clean_val(row.get("final normalized disease id"))
            if not drug_id or not dis_id:
                continue
            # Skip error/invalid drug IDs
            if drug_id.startswith("[") or drug_id == "Error":
                continue

            record = {
                "disease_id": dis_id,
                "disease_label": _clean_val(row.get("final normalized disease label")),
                "sources": {
                    "fda": bool(row.get("FDA")),
                    "ema": bool(row.get("EMA")),
                    "pmda": bool(row.get("PMDA")),
                },
            }
            indications[drug_id].append(record)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = [str(c) for c in rows[0]]
        drug_col = header.index("final normalized drug id")
        dis_id_col = header.index("final normalized disease id")
        dis_label_col = header.index("final normalized disease label")
        fda_col = header.index("FDA")
        ema_col = header.index("EMA")
        pmda_col = header.index("PMDA")

        for row in rows[1:]:
            drug_id = _clean_val(row[drug_col])
            dis_id = _clean_val(row[dis_id_col])
            if not drug_id or not dis_id:
                continue
            if drug_id.startswith("[") or drug_id == "Error":
                continue
            record = {
                "disease_id": dis_id,
                "disease_label": _clean_val(row[dis_label_col]),
                "sources": {
                    "fda": bool(row[fda_col]),
                    "ema": bool(row[ema_col]),
                    "pmda": bool(row[pmda_col]),
                },
            }
            indications[drug_id].append(record)

    return indications


# ---------------------------------------------------------------------------
# Phase 3: Load contraindication list
# ---------------------------------------------------------------------------

def _load_contraindications(data_folder):
    """
    Load contraindicationList.xlsx. Returns dict: drug_curie -> list of contraindication records.
    """
    contraindications = defaultdict(list)
    fpath = os.path.join(data_folder, "contraindicationList.xlsx")
    if not os.path.exists(fpath):
        return contraindications

    if HAS_PANDAS:
        df = pd.read_excel(fpath, engine="openpyxl")
        df = df.where(pd.notnull(df), None)
        for _, row in df.iterrows():
            drug_id = _clean_val(row.get("final normalized drug id"))
            dis_id = _clean_val(row.get("final normalized disease id"))
            if not drug_id or not dis_id:
                continue
            if drug_id.startswith("[") or drug_id == "Error":
                continue

            record = {
                "disease_id": dis_id,
                "disease_label": _clean_val(row.get("final normalized disease label")),
                "is_allergen": _to_bool(str(row.get("is_allergen", "")).strip()),
                "is_diagnostic_agent": _to_bool(str(row.get("is_diagnostic_agent", "")).strip()),
            }
            contraindications[drug_id].append(record)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = [str(c) for c in rows[0]]
        drug_col = header.index("final normalized drug id")
        dis_id_col = header.index("final normalized disease id")
        dis_label_col = header.index("final normalized disease label")
        allergen_col = header.index("is_allergen")
        diag_col = header.index("is_diagnostic_agent")

        for row in rows[1:]:
            drug_id = _clean_val(row[drug_col])
            dis_id = _clean_val(row[dis_id_col])
            if not drug_id or not dis_id:
                continue
            if drug_id.startswith("[") or drug_id == "Error":
                continue
            record = {
                "disease_id": dis_id,
                "disease_label": _clean_val(row[dis_label_col]),
                "is_allergen": _to_bool(str(row[allergen_col]).strip()),
                "is_diagnostic_agent": _to_bool(str(row[diag_col]).strip()),
            }
            contraindications[drug_id].append(record)

    return contraindications


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def load_data(data_folder):
    """
    Parse MeDIC data files and yield BioThings-compatible documents.

    _id = drug CURIE from drugList.tsv (e.g. CHEBI:8327, UNII:N0A21N6RAU).
    All MeDIC data nested under top-level 'medic' key.
    One document per unique drug CURIE.
    """
    drugs = _load_drug_list(data_folder)
    indications = _load_indications(data_folder)
    contraindications = _load_contraindications(data_folder)

    # Build union of all drug IDs (some drugs may have indications/contraindications
    # without appearing in the drug list if released separately)
    all_drug_ids = set(drugs.keys()) | set(indications.keys()) | set(contraindications.keys())

    seen_ids = set()

    for drug_id in all_drug_ids:
        if drug_id in seen_ids:
            continue
        seen_ids.add(drug_id)

        drug_meta = drugs.get(drug_id, {"drug_id": drug_id})

        ind_list = indications.get(drug_id)
        ci_list = contraindications.get(drug_id)

        if ind_list:
            drug_meta["indications"] = ind_list
        if ci_list:
            drug_meta["contraindications"] = ci_list

        # Skip drugs with no indications/contraindications AND no metadata value
        if not ind_list and not ci_list and drug_id not in drugs:
            continue

        doc = {
            "_id": drug_id,
            "medic": drug_meta,
        }

        doc = dict_sweep(unlist(doc), [None])
        yield doc
